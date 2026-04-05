"""
MF Fund Ranker — DUAL ENGINE Model (FIXED VERSION)
====================================
Reads your dashboard_data.xlsx, applies a dual-layer scoring system,
and outputs a ranked Excel file: mf_ranked_screener.xlsx

FIXES APPLIED:
  ✓ return_730d column now detected as 2-year return
  ✓ CAGR conversion handles percentage-based returns correctly
  ✓ Assumptions Summary sheet auto-generated for transparency

DUAL ENGINE ARCHITECTURE:
  ENGINE 1 - LONG-TERM QUALITY (Conservative consistency)
    • Filters: 2Y CAGR > 10%, 3Y CAGR > 12%
    • Scores: 1Y (25%), 2Y CAGR (30%), 3Y CAGR (45%)
    • Focus: Structural soundness & proven compounding
    
  ENGINE 2 - SHORT-TERM MOMENTUM (Tactical trend-following)
    • Scores: 6M (30%), 3M (20%), 1Y (25%), 1M (25%)
    • Focus: Current performance & recovery trajectory
    • Trend confirmation: +5 bonus if 6M > 3M > 1M
    
  COMPOSITE SCORE (Blended)
    • 55% Momentum (Engine 2) + 45% Quality (Engine 1)
    • Only qualifies funds passing quality filters
    • Ranked within category via percentile scoring

Usage:
  pip install pandas openpyxl
  python mf_fund_ranker_dual_engine_FIXED.py

Place dashboard_data.xlsx in the same folder as this script.
Output: mf_ranked_screener.xlsx (with Engine 1 & Engine 2 scores visible + Assumptions sheet)
"""

import pandas as pd
import os
from datetime import datetime
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

# ── CONFIG ──────────────────────────────────────────────────────────────────
INPUT_FILE  = "dashboard_data.xlsx"
OUTPUT_FILE = "mf_ranked_screener.xlsx"

# ── QUALITY FILTER THRESHOLDS (NEW - Dual Engine Model) ──────────────────
# NOTE: 2Y filter disabled if 2Y return column not available
QUALITY_FILTERS = {
    "cagr_2y_min": 0.10,      # 2Y CAGR must be > 10% (if column exists)
    "cagr_3y_min": 0.12,      # 3Y CAGR must be > 12%
}

# ── ENGINE WEIGHTS (NEW - Dual Engine Model) ──────────────────────────────
# ENGINE 1: Long-Term Quality Scoring (within qualified funds)
ENGINE1_WEIGHTS = {
    "return_1y":  0.25,   # 1-year return
    "return_2y":  0.30,   # 2-year CAGR (converted from 2Y return)
    "return_3y":  0.45,   # 3-year CAGR (long-term emphasis)
}

# ENGINE 2: Short-Term Momentum Scoring
ENGINE2_WEIGHTS = {
    "return_6m":  0.30,   # 6-month momentum
    "return_3m":  0.20,   # 3-month momentum
    "return_1y":  0.25,   # 1-year bridge
    "return_1m":  0.25,   # 1-month latest (if available)
}

# ── COMPOSITE BLEND (NEW - Dual Engine Model) ──────────────────────────
COMPOSITE_BLEND = {
    "engine1_quality": 0.45,   # Long-term quality weight
    "engine2_momentum": 0.55,  # Short-term momentum weight (higher = trend-focused)
}

# ── TREND CONFIRMATION BONUS (NEW - Dual Engine Model) ──────────────────
TREND_BONUS = 5.0  # +5 points if 6M > 3M > 1M (confirmed uptrend)

# ── FILTERS ─────────────────────────────────────────────────────────────────
# Set a value to None to skip that filter entirely.
# Column names must match exactly what is in your Excel file (case-insensitive).
FILTERS = {
    "cat_level_1": "Open Ended Schemes",
    "cat_level_2": "Equity Scheme",
    "plan_type":   "Regular",
    "option_type": "Growth",
}

# ── COLUMN AUTO-DETECTION ───────────────────────────────────────────────────
COLUMN_ALIASES = {
    "scheme_name": [
        "scheme_name", "scheme name", "fund name", "name", "schemename",
        "fund", "scheme"
    ],
    "category": [
        "cat_level_3", "cat_level_2", "cat_level_1",
        "category", "scheme_category", "scheme category", "cat",
        "fund_category", "fund category", "sub_category", "sub category",
        "category_name", "sub_type"
    ],
    "amc": [
        "amc", "amc_name", "amc name", "fund_house", "fund house",
        "amcname", "asset_management_company", "house"
    ],
    "return_1m": [
        "return_30d", "return_1m", "1m_return", "1month_return", "ret_1m",
        "trailing_1m", "1m return", "1 month return"
    ],
    "return_3m": [
        "return_90d", "return_3m", "3m_return", "3month_return", "ret_3m",
        "trailing_3m", "3m return", "3 month return"
    ],
    "return_6m": [
        "return_180d", "return_6m", "6m_return", "6month_return", "ret_6m",
        "trailing_6m", "6m return", "6 month return"
    ],
    "return_1y": [
        "return_365d", "return_1y", "1y_return", "1yr_return", "returns_1y",
        "1y return", "1 year return", "1yr", "ret_1y", "cagr_1y", "1y_cagr",
        "trailing_1y", "1year_return", "ann_return_1y"
    ],
    "return_2y": [
        "return_730d", "return_2y", "2y_return", "2yr_return", "returns_2y",
        "2y return", "2 year return", "2yr", "ret_2y", "cagr_2y", "2y_cagr",
        "trailing_2y", "2year_return", "ann_return_2y"
    ],
    "return_3y": [
        "return_1095d", "return_3y", "3y_return", "3yr_return", "returns_3y",
        "3y return", "3 year return", "3yr", "ret_3y", "cagr_3y", "3y_cagr",
        "trailing_3y", "3year_return", "ann_return_3y"
    ],
    "nav": [
        "nav", "net_asset_value", "current_nav", "latest_nav"
    ],
    "plan": [
        "plan", "plan_type", "scheme_plan", "direct_regular"
    ],
}

COLORS = {
    "header_bg":   "1A3A5C",   # dark navy
    "header_fg":   "FFFFFF",
    "rank1_bg":    "FFD700",   # gold
    "rank2_bg":    "E8E8E8",   # silver
    "rank3_bg":    "D4956A",   # bronze
    "positive":    "1E7A4B",   # dark green
    "negative":    "C0392B",   # red
    "cat_header":  "2D6A8A",   # teal
    "cat_fg":      "FFFFFF",
    "alt_row":     "F5F0E8",   # cream
    "score_bg":    "EBF5FB",
    "border":      "CCCCCC",
    "engine1_bg":  "E3F2FD",   # Light blue (Quality)
    "engine2_bg":  "FFF3E0",   # Light amber (Momentum)
    "assumption_header": "2D5016",  # dark green
}


def detect_column(df_cols, key):
    """Find the actual column name from aliases."""
    cols_lower = {c.lower().strip(): c for c in df_cols}
    for alias in COLUMN_ALIASES.get(key, []):
        if alias.lower() in cols_lower:
            return cols_lower[alias.lower()]
    return None


def apply_filters(df):
    """
    Apply FILTERS to the DataFrame.
    Each key in FILTERS must match a column name in the data (case-insensitive).
    Values are matched case-insensitively; set a filter value to None to skip it.
    """
    cols_lower = {c.lower().strip(): c for c in df.columns}
    active_filters = {k: v for k, v in FILTERS.items() if v is not None}

    if not active_filters:
        print("   ℹ️  No filters configured — using all rows.")
        return df

    print(f"\n🔎 Applying filters ({len(active_filters)} active):")
    for col_key, value in active_filters.items():
        actual_col = cols_lower.get(col_key.lower().strip())
        if actual_col is None:
            print(f"   ⚠️  Filter column '{col_key}' not found in data — skipping this filter.")
            continue
        before = len(df)
        df = df[df[actual_col].astype(str).str.strip().str.lower() == str(value).strip().lower()]
        print(f"   ✅ {col_key} = '{value}'  →  {before} → {len(df)} rows")

    print(f"   📋 Rows after all filters: {len(df)}")
    return df


def load_data(filepath):
    """Load all sheets, combine, apply filters, and return a unified DataFrame."""
    print(f"\n📂 Reading: {filepath}")
    sheets = pd.read_excel(filepath, sheet_name=None)
    print(f"   Sheets found: {list(sheets.keys())}")

    frames = []
    for name, df in sheets.items():
        df["_source_sheet"] = name
        frames.append(df)

    combined = pd.concat(frames, ignore_index=True)
    print(f"   Total rows across all sheets: {len(combined)}")
    print(f"   Columns: {list(combined.columns)}")

    combined = apply_filters(combined)
    return combined


def map_columns(df):
    """Detect and map standard field names."""
    mapping = {}
    for key in COLUMN_ALIASES:
        col = detect_column(df.columns, key)
        mapping[key] = col
        status = f"✅ '{col}'" if col else "❌ not found"
        print(f"   {key:<15} → {status}")
    return mapping


def to_numeric(series):
    """Clean and convert a series to numeric, coercing errors."""
    if series is None:
        return pd.Series(dtype=float)
    s = series.astype(str).str.replace('%', '').str.replace(',', '').str.strip()
    return pd.to_numeric(s, errors='coerce')


def percentile_score(series):
    """Rank-based percentile score 0–100 within the group."""
    ranks = series.rank(method='min', na_option='bottom')
    return (ranks - 1) / max(len(series) - 1, 1) * 100


# ════════════════════════════════════════════════════════════════════════════
# NEW FUNCTIONS - DUAL ENGINE MODEL (WITH FIXES)
# ════════════════════════════════════════════════════════════════════════════

def convert_2y_to_cagr(return_2y):
    """
    Convert 2-year total return to annualized CAGR.
    
    FIXED: Now correctly handles returns in percentage form (e.g., 12.22 = 12.22%)
    Formula: CAGR = (1 + return)^(1/years) - 1
    For 2 years: CAGR = sqrt(1 + return) - 1
    
    Args:
        return_2y: Total 2-year return (in % form, e.g., 12.22 for +12.22%)
    
    Returns:
        Annualized CAGR (in % form), or NaN if input is NaN
    """
    if pd.isna(return_2y):
        return float('nan')
    try:
        # Ensure return_2y is numeric
        r = float(return_2y)
        
        # Data comes in percentage form (e.g., 12.22 = 12.22%)
        # Convert to decimal: 12.22% → 0.1222
        r_decimal = r / 100.0
        
        # CAGR = (1 + r)^(1/2) - 1
        cagr = (1 + r_decimal) ** (1/2) - 1
        
        # Convert back to percentage form
        cagr_pct = cagr * 100.0
        
        return cagr_pct
    except (ValueError, TypeError):
        return float('nan')


def score_engine1_quality(df, mapping):
    """
    ENGINE 1: LONG-TERM QUALITY SCORING
    
    Purpose: Identify structurally sound funds with proven 2Y & 3Y performance
    
    Filters (hard gates):
      • 2Y CAGR > 10%
      • 3Y CAGR > 12%
    
    Scoring (within qualified funds):
      • 1Y Return (25%)
      • 2Y CAGR (30%)
      • 3Y CAGR (45%)
    
    Returns: DataFrame with _engine1_score column (0-100 per category)
    """
    print("\n⚙️  ENGINE 1: Long-Term Quality Scoring...")
    
    df = df.copy()
    
    # Extract and convert to numeric
    r1y = to_numeric(df[mapping["return_1y"]] if mapping["return_1y"] else None)
    r2y_raw = to_numeric(df[mapping["return_2y"]] if mapping["return_2y"] else None)
    r3y = to_numeric(df[mapping["return_3y"]] if mapping["return_3y"] else None)
    
    # Convert 2Y return to CAGR
    r2y_cagr = r2y_raw.apply(convert_2y_to_cagr)
    
    df["_r1y"] = r1y
    df["_r2y_cagr"] = r2y_cagr
    df["_r3y"] = r3y
    
    # ── QUALITY FILTERS (Hard Gates) ──────────────────────────────────────
    print(f"   Applying quality filters...")
    before_filter = len(df)
    
    # Filter 1: 2Y CAGR > 10% (only if 2Y column exists)
    if mapping["return_2y"] is not None:
        # Note: r2y_cagr is now in percentage form (e.g., 6.1 for 6.1%)
        # QUALITY_FILTERS["cagr_2y_min"] is decimal (0.10 = 10%)
        # Convert to percentage: 0.10 * 100 = 10
        mask_2y = df["_r2y_cagr"] > (QUALITY_FILTERS["cagr_2y_min"] * 100)
        print(f"     2Y CAGR > {QUALITY_FILTERS['cagr_2y_min']*100}%: {mask_2y.sum()}/{before_filter} funds qualify")
    else:
        print(f"     2Y CAGR > {QUALITY_FILTERS['cagr_2y_min']*100}%: ⚠️  Column not found (skipping this filter)")
        mask_2y = pd.Series(True, index=df.index)  # Don't filter on missing column
    
    # Filter 2: 3Y CAGR > 12% (3Y return is also in percentage form)
    # Need to convert decimal filter to percentage
    mask_3y = df["_r3y"] > (QUALITY_FILTERS["cagr_3y_min"] * 100)
    print(f"     3Y CAGR > {QUALITY_FILTERS['cagr_3y_min']*100}%: {mask_3y.sum()}/{before_filter} funds qualify")
    
    # Combined mask
    quality_mask = mask_2y & mask_3y
    df["_qualifies_quality"] = quality_mask
    qualified_count = quality_mask.sum()
    print(f"   ✅ Both filters: {qualified_count}/{before_filter} funds qualify for Engine 1")
    
    # ── SCORING (Percentile within category for qualified funds) ───────────
    cat_col = mapping["category"]
    if not cat_col:
        df["_category_clean"] = "All Funds"
    else:
        df["_category_clean"] = df[cat_col].astype(str).str.strip().str.title()
    
    df["_engine1_score"] = 0.0
    
    # Determine which columns are available
    cols_available = {
        "return_1y":  mapping["return_1y"] is not None,
        "return_2y":  mapping["return_2y"] is not None,
        "return_3y":  mapping["return_3y"] is not None,
    }
    available_count = sum(cols_available.values())
    
    if available_count == 0:
        print(f"   ⚠️  No return columns available for Engine 1 — setting score to 0")
        return df
    
    # Recalculate weights if columns are missing
    weights = ENGINE1_WEIGHTS.copy()
    if not cols_available["return_1y"]:
        weights.pop("return_1y", None)
    if not cols_available["return_2y"]:
        weights.pop("return_2y", None)
    if not cols_available["return_3y"]:
        weights.pop("return_3y", None)
    
    # Normalize weights
    total_weight = sum(weights.values())
    weights = {k: v / total_weight for k, v in weights.items()}
    
    print(f"   Scoring metrics (Engine 1):")
    for metric, w in weights.items():
        print(f"     {metric}: {w*100:.0f}%")
    
    # Score per category
    categories = df["_category_clean"].unique()
    for cat in categories:
        cat_mask = df["_category_clean"] == cat
        cat_qualified = cat_mask & quality_mask
        
        if cat_qualified.sum() == 0:
            continue  # No qualified funds in this category
        
        # Initialize composite score (for qualified funds only)
        composite_scores = pd.Series(0.0, index=df[cat_qualified].index)
        
        # Score each available metric with percentile ranking
        if "return_1y" in weights:
            r1y_scores = percentile_score(df.loc[cat_qualified, "_r1y"])
            composite_scores += r1y_scores * weights["return_1y"]
        
        if "return_2y" in weights:
            r2y_scores = percentile_score(df.loc[cat_qualified, "_r2y_cagr"])
            composite_scores += r2y_scores * weights["return_2y"]
        
        if "return_3y" in weights:
            r3y_scores = percentile_score(df.loc[cat_qualified, "_r3y"])
            composite_scores += r3y_scores * weights["return_3y"]
        
        df.loc[cat_qualified, "_engine1_score"] = composite_scores
    
    return df


def score_engine2_momentum(df, mapping):
    """
    ENGINE 2: SHORT-TERM MOMENTUM SCORING
    
    Purpose: Capture recent trends and recovery trajectory
    
    Scoring:
      • 6M Return (30%)
      • 3M Return (20%)
      • 1Y Return (25%)
      • 1M Return (25%)
    
    Bonus:
      • +5 points if 6M > 3M > 1M (confirmed uptrend)
    
    Returns: DataFrame with _engine2_score column (0-100 per category)
    """
    print("\n⚙️  ENGINE 2: Short-Term Momentum Scoring...")
    
    df = df.copy()
    
    # Extract and convert to numeric
    r1m = to_numeric(df[mapping["return_1m"]] if mapping["return_1m"] else None)
    r3m = to_numeric(df[mapping["return_3m"]] if mapping["return_3m"] else None)
    r6m = to_numeric(df[mapping["return_6m"]] if mapping["return_6m"] else None)
    r1y = to_numeric(df[mapping["return_1y"]] if mapping["return_1y"] else None)
    
    df["_r1m"] = r1m
    df["_r3m"] = r3m
    df["_r6m"] = r6m
    df["_r1y_engine2"] = r1y
    
    # Determine which columns are available
    cols_available = {
        "return_6m":  mapping["return_6m"] is not None,
        "return_3m":  mapping["return_3m"] is not None,
        "return_1y":  mapping["return_1y"] is not None,
        "return_1m":  mapping["return_1m"] is not None,
    }
    available_count = sum(cols_available.values())
    
    if available_count == 0:
        print(f"   ⚠️  No return columns available for Engine 2 — setting score to 0")
        df["_engine2_score"] = 0.0
        return df
    
    # Recalculate weights if columns are missing
    weights = ENGINE2_WEIGHTS.copy()
    if not cols_available["return_6m"]:
        weights.pop("return_6m", None)
    if not cols_available["return_3m"]:
        weights.pop("return_3m", None)
    if not cols_available["return_1y"]:
        weights.pop("return_1y", None)
    if not cols_available["return_1m"]:
        weights.pop("return_1m", None)
    
    # Normalize weights
    total_weight = sum(weights.values())
    weights = {k: v / total_weight for k, v in weights.items()}
    
    print(f"   Scoring metrics (Engine 2):")
    for metric, w in weights.items():
        print(f"     {metric}: {w*100:.0f}%")
    
    # Trend bonus logic
    df["_trend_signal"] = ""
    trend_mask = (df["_r6m"] > df["_r3m"]) & (df["_r3m"] > df["_r1m"])
    df.loc[trend_mask, "_trend_signal"] = "📈 Uptrend"
    print(f"   Trend bonus: {trend_mask.sum()} funds with confirmed 6M > 3M > 1M")
    
    df["_engine2_score"] = 0.0
    
    # Score per category
    categories = df["_category_clean"].unique()
    for cat in categories:
        cat_mask = df["_category_clean"] == cat
        
        # Initialize composite score
        composite_scores = pd.Series(0.0, index=df[cat_mask].index)
        
        # Score each available metric with percentile ranking
        if "return_6m" in weights:
            r6m_scores = percentile_score(df.loc[cat_mask, "_r6m"])
            composite_scores += r6m_scores * weights["return_6m"]
        
        if "return_3m" in weights:
            r3m_scores = percentile_score(df.loc[cat_mask, "_r3m"])
            composite_scores += r3m_scores * weights["return_3m"]
        
        if "return_1y" in weights:
            r1y_scores = percentile_score(df.loc[cat_mask, "_r1y_engine2"])
            composite_scores += r1y_scores * weights["return_1y"]
        
        if "return_1m" in weights:
            r1m_scores = percentile_score(df.loc[cat_mask, "_r1m"])
            composite_scores += r1m_scores * weights["return_1m"]
        
        # Add trend bonus for confirmed uptrend
        uptrend = df.loc[cat_mask, "_trend_signal"] == "📈 Uptrend"
        composite_scores.loc[uptrend.values] += TREND_BONUS
        
        # Cap at 100
        composite_scores = composite_scores.clip(0, 100)
        
        df.loc[cat_mask, "_engine2_score"] = composite_scores
    
    return df


def score_funds(df, mapping):
    """Run both engine scorers and blend into composite."""
    df = score_engine1_quality(df, mapping)
    df = score_engine2_momentum(df, mapping)
    
    print("\n⚙️  ENGINE BLEND: Computing Composite Score...")
    
    # Compute composite score
    df["_composite_score"] = (
        df["_engine1_score"] * COMPOSITE_BLEND["engine1_quality"] +
        df["_engine2_score"] * COMPOSITE_BLEND["engine2_momentum"]
    )
    
    # Rank within each category
    df["_rank"] = df.groupby("_category_clean")["_composite_score"].rank(
        method='min', ascending=False
    ).astype(int)
    
    return df


def get_col_val(row, col_name):
    """Safely get column value."""
    if col_name and col_name in row.index:
        return row[col_name]
    return None


def pct(val):
    """Format numeric value as percentage string."""
    if pd.isna(val) or val is None:
        return "—"
    try:
        return f"{float(val):+.2f}%"
    except:
        return "—"


def score_color(val):
    """Return color hex for score value."""
    if not isinstance(val, (int, float)):
        val = float(str(val).replace("%", "")) if val else 0
    if val >= 75:
        return "1E7A4B"
    elif val >= 50:
        return "F39C12"
    else:
        return "C0392B"


def cell_font(bold=False, size=9):
    """Create a cell font."""
    return Font(name="Arial", bold=bold, size=size, color="000000")


def hdr_font():
    """Create header font."""
    return Font(name="Arial", bold=True, size=10, color=COLORS["header_fg"])


def fill(color_hex):
    """Create a fill with given hex color."""
    return PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")


import re

def clean_sheet_name(name):
    name = re.sub(r'[\\/*?:\[\]]', '', str(name))  # remove invalid Excel chars
    return name[:31]  # Excel sheet name limit


def build_excel(df_scored, mapping, output_path):
    """Build Excel workbook with category sheets + summary + assumptions."""
    from openpyxl import Workbook
    
    print("\n📝 Building Excel workbook...")
    wb = Workbook()
    
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    border = Border(
        left=Side(style='thin', color=COLORS["border"]),
        right=Side(style='thin', color=COLORS["border"]),
        top=Side(style='thin', color=COLORS["border"]),
        bottom=Side(style='thin', color=COLORS["border"])
    )
    
    RETURN_COLS = {4, 5, 6, 7, 8}  # Columns for returns (1M, 3M, 6M, 1Y, 3Y)
    
    categories = sorted(df_scored["_category_clean"].unique())
    
    # ── SUMMARY SHEET ────────────────────────────────────────────────────
    ws_sum = wb.create_sheet(title="🏆 SUMMARY", index=0)
    
    ws_sum.merge_cells("A1:K1")
    ws_sum["A1"] = "MF INTELLIGENCE — DUAL ENGINE TOP FUND PER CATEGORY"
    ws_sum["A1"].font = Font(name="Arial", bold=True, size=14, color="FFFFFF")
    ws_sum["A1"].fill = fill("0D1117")
    ws_sum["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_sum.row_dimensions[1].height = 30
    
    ws_sum.merge_cells("A2:K2")
    ws_sum["A2"] = (
        "Dual Engine: Long-Term Quality (45% weight, filters 2Y CAGR>10% & 3Y CAGR>12%) "
        "+ Short-Term Momentum (55% weight, 6M/3M/1M returns)  |  Category: cat_level_3"
    )
    ws_sum["A2"].font = Font(name="Arial", italic=True, size=8, color="555555")
    ws_sum["A2"].fill = fill("F5F5F5")
    ws_sum["A2"].alignment = Alignment(horizontal="center")
    ws_sum.row_dimensions[2].height = 16
    
    sum_headers = [
        "Category", "#1 Fund", "AMC",
        "6M Return", "3M Return", "1M Return",
        "1Y Return", "3Y CAGR",
        "Engine 1 (Quality)", "Engine 2 (Momentum)", "Signal"
    ]
    for ci, hdr in enumerate(sum_headers, 1):
        c = ws_sum.cell(row=3, column=ci, value=hdr)
        c.font = hdr_font()
        c.fill = fill(COLORS["header_bg"])
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border
    ws_sum.row_dimensions[3].height = 18
    
    SUM_RETURN_COLS = {4, 5, 6, 7, 8}
    SUM_ENGINE_COLS  = {9, 10}
    
    for i, cat in enumerate(categories, 4):
        cat_df = df_scored[df_scored["_category_clean"] == cat].sort_values("_rank")
        if cat_df.empty: 
            continue
        
        top   = cat_df.iloc[0]
        name  = get_col_val(top, mapping["scheme_name"]) or "—"
        amc   = get_col_val(top, mapping["amc"])         or "—"
        r1m   = top.get("_r1m")
        r3m   = top.get("_r3m")
        r6m   = top.get("_r6m")
        r1y   = top.get("_r1y")
        r3y   = top.get("_r3y")
        
        # NEW: Engine scores
        engine1 = top.get("_engine1_score", 0)
        engine2 = top.get("_engine2_score", 0)
        composite = top.get("_composite_score", 0)
        
        signal = "⭐ Strong Buy" if composite >= 75 else "✅ Buy" if composite >= 55 else "⚠️ Watch"
        
        row_bg = COLORS["alt_row"] if i % 2 == 0 else "FFFFFF"
        vals = [cat, name, amc, pct(r6m), pct(r3m), pct(r1m), pct(r1y), pct(r3y), 
                round(engine1, 1), round(engine2, 1), signal]
        
        for ci, v in enumerate(vals, 1):
            c = ws_sum.cell(row=i, column=ci, value=v)
            c.font = cell_font(size=9)
            c.fill = fill(row_bg)
            c.border = border
            c.alignment = Alignment(
                horizontal="center" if ci != 2 else "left",
                vertical="center", wrap_text=(ci == 2)
            )
            
            # Color returns
            if ci in SUM_RETURN_COLS and isinstance(v, str) and v != "—":
                num = float(v.replace('%', ''))
                c.font = Font(
                    name="Arial", size=9,
                    color=COLORS["positive"] if num > 0 else COLORS["negative"]
                )
            
            # Color engine scores
            if ci in SUM_ENGINE_COLS:
                c.font = Font(name="Arial", bold=True, size=9, color=score_color(v))
                if ci == 9:
                    c.fill = fill(COLORS["engine1_bg"])
                elif ci == 10:
                    c.fill = fill(COLORS["engine2_bg"])
        
        ws_sum.row_dimensions[i].height = 18
    
    # Summary column widths
    for ci, w in enumerate([28, 52, 20, 11, 11, 11, 11, 11, 16, 16, 14], 1):
        ws_sum.column_dimensions[get_column_letter(ci)].width = w
    
    ws_sum.freeze_panes = "A4"
    
    # ── ASSUMPTIONS SUMMARY SHEET ────────────────────────────────────────
    ws_assumptions = wb.create_sheet(title="📋 ASSUMPTIONS", index=1)
    
    # Title
    ws_assumptions.merge_cells("A1:D1")
    ws_assumptions["A1"] = "DUAL ENGINE MODEL — ASSUMPTIONS & METHODOLOGY"
    ws_assumptions["A1"].font = Font(name="Arial", bold=True, size=14, color="FFFFFF")
    ws_assumptions["A1"].fill = fill(COLORS["assumption_header"])
    ws_assumptions["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_assumptions.row_dimensions[1].height = 25
    
    # Subheader
    ws_assumptions.merge_cells("A2:D2")
    ws_assumptions["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws_assumptions["A2"].font = Font(name="Arial", italic=True, size=9, color="555555")
    ws_assumptions["A2"].fill = fill("F5F5F5")
    ws_assumptions.row_dimensions[2].height = 14
    
    row = 4
    
    # ── ENGINE WEIGHTS ──────────────────────────────────────────────────
    ws_assumptions[f"A{row}"] = "ENGINE 1 & ENGINE 2 WEIGHTS"
    ws_assumptions[f"A{row}"].font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    ws_assumptions[f"A{row}"].fill = fill("2D6A8A")
    ws_assumptions.merge_cells(f"A{row}:D{row}")
    ws_assumptions.row_dimensions[row].height = 18
    row += 1
    
    ws_assumptions[f"A{row}"] = "COMPOSITE BLEND RATIO"
    ws_assumptions[f"B{row}"] = f"{COMPOSITE_BLEND['engine1_quality']*100:.0f}% (Engine 1)"
    ws_assumptions[f"C{row}"] = f"{COMPOSITE_BLEND['engine2_momentum']*100:.0f}% (Engine 2)"
    for col in ["A", "B", "C"]:
        ws_assumptions[f"{col}{row}"].font = Font(name="Arial", bold=True, size=10)
        ws_assumptions[f"{col}{row}"].fill = fill("E3F2FD")
    row += 2
    
    ws_assumptions[f"A{row}"] = "ENGINE 1: LONG-TERM QUALITY (45%)"
    ws_assumptions[f"A{row}"].font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    ws_assumptions[f"A{row}"].fill = fill("1A3A5C")
    ws_assumptions.merge_cells(f"A{row}:D{row}")
    ws_assumptions.row_dimensions[row].height = 16
    row += 1
    
    ws_assumptions[f"A{row}"] = "Metric"
    ws_assumptions[f"B{row}"] = "Weight"
    for col in ["A", "B"]:
        ws_assumptions[f"{col}{row}"].font = Font(name="Arial", bold=True, size=9, color="FFFFFF")
        ws_assumptions[f"{col}{row}"].fill = fill("E3F2FD")
    row += 1
    
    for metric, weight in ENGINE1_WEIGHTS.items():
        ws_assumptions[f"A{row}"] = metric.replace("return_", "").upper()
        ws_assumptions[f"B{row}"] = f"{weight*100:.0f}%"
        ws_assumptions.row_dimensions[row].height = 14
        row += 1
    
    row += 1
    ws_assumptions[f"A{row}"] = "ENGINE 1: QUALITY FILTERS (Hard Gates)"
    ws_assumptions[f"A{row}"].font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    ws_assumptions[f"A{row}"].fill = fill("1A3A5C")
    ws_assumptions.merge_cells(f"A{row}:D{row}")
    ws_assumptions.row_dimensions[row].height = 16
    row += 1
    
    ws_assumptions[f"A{row}"] = "Filter"
    ws_assumptions[f"B{row}"] = "Threshold"
    for col in ["A", "B"]:
        ws_assumptions[f"{col}{row}"].font = Font(name="Arial", bold=True, size=9, color="FFFFFF")
        ws_assumptions[f"{col}{row}"].fill = fill("E3F2FD")
    row += 1
    
    ws_assumptions[f"A{row}"] = "2Y CAGR Minimum"
    ws_assumptions[f"B{row}"] = f">{QUALITY_FILTERS['cagr_2y_min']*100:.0f}%"
    ws_assumptions.row_dimensions[row].height = 14
    row += 1
    
    ws_assumptions[f"A{row}"] = "3Y CAGR Minimum"
    ws_assumptions[f"B{row}"] = f">{QUALITY_FILTERS['cagr_3y_min']*100:.0f}%"
    ws_assumptions.row_dimensions[row].height = 14
    row += 3
    
    ws_assumptions[f"A{row}"] = "ENGINE 2: SHORT-TERM MOMENTUM (55%)"
    ws_assumptions[f"A{row}"].font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    ws_assumptions[f"A{row}"].fill = fill("1A3A5C")
    ws_assumptions.merge_cells(f"A{row}:D{row}")
    ws_assumptions.row_dimensions[row].height = 16
    row += 1
    
    ws_assumptions[f"A{row}"] = "Metric"
    ws_assumptions[f"B{row}"] = "Weight"
    for col in ["A", "B"]:
        ws_assumptions[f"{col}{row}"].font = Font(name="Arial", bold=True, size=9, color="FFFFFF")
        ws_assumptions[f"{col}{row}"].fill = fill("FFF3E0")
    row += 1
    
    for metric, weight in ENGINE2_WEIGHTS.items():
        ws_assumptions[f"A{row}"] = metric.replace("return_", "").upper()
        ws_assumptions[f"B{row}"] = f"{weight*100:.0f}%"
        ws_assumptions.row_dimensions[row].height = 14
        row += 1
    
    row += 1
    ws_assumptions[f"A{row}"] = "TREND CONFIRMATION BONUS"
    ws_assumptions[f"A{row}"].font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    ws_assumptions[f"A{row}"].fill = fill("1A3A5C")
    ws_assumptions.merge_cells(f"A{row}:D{row}")
    ws_assumptions.row_dimensions[row].height = 16
    row += 1
    
    ws_assumptions[f"A{row}"] = "Condition"
    ws_assumptions[f"B{row}"] = "Bonus Points"
    for col in ["A", "B"]:
        ws_assumptions[f"{col}{row}"].font = Font(name="Arial", bold=True, size=9, color="FFFFFF")
        ws_assumptions[f"{col}{row}"].fill = fill("FFF3E0")
    row += 1
    
    ws_assumptions[f"A{row}"] = "6M Return > 3M Return > 1M Return"
    ws_assumptions[f"B{row}"] = f"+{TREND_BONUS:.0f}"
    ws_assumptions.row_dimensions[row].height = 14
    row += 3
    
    ws_assumptions[f"A{row}"] = "RANKING LOGIC"
    ws_assumptions[f"A{row}"].font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    ws_assumptions[f"A{row}"].fill = fill("1A3A5C")
    ws_assumptions.merge_cells(f"A{row}:D{row}")
    ws_assumptions.row_dimensions[row].height = 16
    row += 1
    
    ws_assumptions[f"A{row}"] = "• Funds ranked within each category by Composite Score"
    ws_assumptions[f"A{row}"].alignment = Alignment(wrap_text=True, vertical="top")
    ws_assumptions.row_dimensions[row].height = 16
    row += 1
    
    ws_assumptions[f"A{row}"] = "• Composite = 45% × Engine 1 + 55% × Engine 2"
    ws_assumptions[f"A{row}"].alignment = Alignment(wrap_text=True, vertical="top")
    ws_assumptions.row_dimensions[row].height = 16
    row += 1
    
    ws_assumptions[f"A{row}"] = "• Only funds passing quality filters get Engine 1 score"
    ws_assumptions[f"A{row}"].alignment = Alignment(wrap_text=True, vertical="top")
    ws_assumptions.row_dimensions[row].height = 16
    row += 1
    
    ws_assumptions[f"A{row}"] = "• All funds get Engine 2 score (no restrictions)"
    ws_assumptions[f"A{row}"].alignment = Alignment(wrap_text=True, vertical="top")
    ws_assumptions.row_dimensions[row].height = 16
    row += 3
    
    ws_assumptions[f"A{row}"] = "SIGNAL INTERPRETATION"
    ws_assumptions[f"A{row}"].font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    ws_assumptions[f"A{row}"].fill = fill("1A3A5C")
    ws_assumptions.merge_cells(f"A{row}:D{row}")
    ws_assumptions.row_dimensions[row].height = 16
    row += 1
    
    signals = [
        ("⭐ Strong Buy", "Composite ≥ 75", "Exceptional quality + momentum"),
        ("✅ Buy", "Composite ≥ 55", "Solid quality + good momentum"),
        ("⚠️ Watch", "Composite < 55", "Below target or failed quality gate"),
    ]
    
    for signal, threshold, desc in signals:
        ws_assumptions[f"A{row}"] = signal
        ws_assumptions[f"B{row}"] = threshold
        ws_assumptions[f"C{row}"] = desc
        ws_assumptions[f"A{row}"].font = Font(name="Arial", bold=True, size=9)
        ws_assumptions.row_dimensions[row].height = 14
        row += 1
    
    # Column widths
    ws_assumptions.column_dimensions["A"].width = 35
    ws_assumptions.column_dimensions["B"].width = 25
    ws_assumptions.column_dimensions["C"].width = 30
    
    # ── CATEGORY SHEETS ──────────────────────────────────────────────────
    for cat in categories:
        cat_df = df_scored[df_scored["_category_clean"] == cat].sort_values("_rank")
        if cat_df.empty:
            continue
        
        # Create sheet (Excel sheet name limit is 31 chars)
        sheet_name = clean_sheet_name(cat)
        ws = wb.create_sheet(title=sheet_name)
        
        # Headers
        ws.merge_cells("A1:K1")
        ws["A1"] = f"🎯 {cat.upper()} — DUAL ENGINE RANKING"
        ws["A1"].font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
        ws["A1"].fill = fill("0D1117")
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 25
        
        # Subheader
        ws.merge_cells("A2:K2")
        ws["A2"] = "Ranked by: 45% Quality (E1) + 55% Momentum (E2) | Refer 📋 ASSUMPTIONS sheet for methodology"
        ws["A2"].font = Font(name="Arial", italic=True, size=8, color="666666")
        ws["A2"].fill = fill("F0F0F0")
        ws["A2"].alignment = Alignment(horizontal="left", wrap_text=True)
        ws.row_dimensions[2].height = 16
        
        headers = [
            "Rank", "Scheme Name", "AMC",
            "1M Return", "3M Return", "6M Return",
            "1Y Return", "3Y CAGR",
            "Engine 1 (Quality)", "Engine 2 (Momentum)", "Composite Score"
        ]
        for ci, hdr in enumerate(headers, 1):
            c = ws.cell(row=4, column=ci, value=hdr)
            c.font = hdr_font()
            c.fill = fill(COLORS["header_bg"])
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = border
        ws.row_dimensions[4].height = 20
        
        # Data rows
        for i, (idx, row_data) in enumerate(cat_df.iterrows(), 5):
            rank = row_data.get("_rank", "—")
            name = get_col_val(row_data, mapping["scheme_name"]) or "—"
            amc  = get_col_val(row_data, mapping["amc"]) or "—"
            
            r1m = row_data.get("_r1m")
            r3m = row_data.get("_r3m")
            r6m = row_data.get("_r6m")
            r1y = row_data.get("_r1y")
            r3y = row_data.get("_r3y")
            
            engine1 = row_data.get("_engine1_score", 0)
            engine2 = row_data.get("_engine2_score", 0)
            composite = row_data.get("_composite_score", 0)
            
            values = [
                rank, name, amc,
                pct(r1m), pct(r3m), pct(r6m),
                pct(r1y), pct(r3y),
                round(engine1, 1), round(engine2, 1), round(composite, 1)
            ]
            
            # Determine row background
            if rank == 1:    row_bg = COLORS["rank1_bg"]
            elif rank == 2:  row_bg = COLORS["rank2_bg"]
            elif rank == 3:  row_bg = COLORS["rank3_bg"]
            elif i % 2 == 0: row_bg = COLORS["alt_row"]
            else:             row_bg = "FFFFFF"
            
            for col_idx, val in enumerate(values, 1):
                cell = ws.cell(row=i, column=col_idx, value=val)
                cell.font = cell_font(bold=(rank <= 3), size=9)
                cell.fill = fill(row_bg)
                cell.border = border
                cell.alignment = Alignment(
                    horizontal="center", vertical="center",
                    wrap_text=(col_idx == 2)
                )
                
                # Color returns (columns 4-8)
                if col_idx in RETURN_COLS and isinstance(val, str) and val != "—":
                    num = float(val.replace('%', ''))
                    cell.font = Font(
                        name="Arial", size=9, bold=(rank <= 3),
                        color=COLORS["positive"] if num > 0 else COLORS["negative"]
                    )
                
                # Color engine scores (columns 9-11)
                if col_idx in {9, 10, 11}:  # Engine 1, Engine 2, Composite
                    cell.font = Font(name="Arial", bold=True, size=9, color=score_color(val))
                    # Light background for engine columns
                    if col_idx == 9:
                        cell.fill = fill(COLORS["engine1_bg"])
                    elif col_idx == 10:
                        cell.fill = fill(COLORS["engine2_bg"])
            
            ws.row_dimensions[i].height = 16
        
        # Column widths (adjusted for new columns)
        widths = [6, 52, 20, 11, 11, 11, 11, 11, 16, 16, 14]
        for ci, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w
        
        ws.freeze_panes = "A5"
        print(f"   ✅ {cat} — {len(cat_df)} funds ranked")
    
    wb.save(output_path)
    print(f"\n✅ Output saved → {output_path}")


def main():
    print("\n" + "="*80)
    print("  MF FUND RANKER — DUAL ENGINE MODEL (FIXED VERSION)")
    print("  Long-Term Quality (45%) + Short-Term Momentum (55%)")
    print("="*80)
    
    if not os.path.exists(INPUT_FILE):
        print(f"\n❌ File not found: {INPUT_FILE}")
        print("   Place dashboard_data.xlsx in the same folder as this script and re-run.")
        return
    
    df = load_data(INPUT_FILE)
    
    if df.empty:
        print("\n❌ No rows remain after applying filters. Check your FILTERS config.")
        return
    
    print("\n🔍 Auto-detecting columns...")
    mapping = map_columns(df)
    
    missing_critical = [k for k in ("scheme_name", "return_1y", "return_3y") if not mapping[k]]
    if missing_critical:
        print(f"\n⚠️  Critical columns not found: {missing_critical}")
        print("   Available columns:", list(df.columns))
        print("   → Update COLUMN_ALIASES at the top of this script to match your column names.")
        return
    
    print("\n⚙️  Scoring funds with dual engine model...")
    df_scored = score_funds(df, mapping)
    
    print("\n📝 Building Excel output (with Assumptions Summary)...")
    build_excel(df_scored, mapping, OUTPUT_FILE)
    
    print("\n" + "="*80)
    print(f"  📊 MF Ranked Screener → {OUTPUT_FILE}")
    print(f"\n  📁 What's inside:")
    print(f"     • 🏆 SUMMARY tab — Top fund per category at a glance")
    print(f"     • 📋 ASSUMPTIONS tab — All parameters & methodology (NEW!)")
    print(f"     • One tab per fund category — All schemes ranked by composite score")
    print(f"\n  🎯 Scoring Model:")
    print(f"     • Engine 1 (Quality): 2Y CAGR > 10%, 3Y CAGR > 12% [1Y/2Y/3Y returns]")
    print(f"     • Engine 2 (Momentum): 6M/3M/1M trend strength [+5 bonus if 6M>3M>1M]")
    print(f"     • Composite: 45% Quality + 55% Momentum (category-wise percentile)")
    print(f"\n  📈 Signals:")
    print(f"     • ⭐ Strong Buy:  Composite Score ≥ 75")
    print(f"     • ✅ Buy:         Composite Score ≥ 55")
    print(f"     • ⚠️  Watch:      Composite Score ≥ 40")
    print(f"     • ❌ Avoid:       Composite Score < 40 (or fails quality filters)")
    print("="*80 + "\n")


if __name__ == "__main__":
    main()
